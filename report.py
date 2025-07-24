import streamlit as st
import requests
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import time
import sys

# Streamlit page configuration
st.set_page_config(
    page_title="Report Processor",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'processed_reports' not in st.session_state:
    st.session_state.processed_reports = []
if 'all_alerts' not in st.session_state:
    st.session_state.all_alerts = []

reports_folder = "doreports"
processed_folder = "poreports"

for folder in [reports_folder, processed_folder]:
    if not os.path.exists(folder):
        os.makedirs(folder)

def parse_tag(tag):
    """
    Parse tag format like "30+DPD @ 75 DOB" to extract DPD and DOB values
    """
    match = re.search(r'(\d+)\+DPD\s*@\s*(\d+)\s*DOB', tag, re.IGNORECASE)
    if match:
        dpd = int(match.group(1))
        dob = int(match.group(2))
        return dpd, dob
    return None, None

def calculate_approved_month(run_month, tag):
    """
    Calculate approved month based on tag formula
    run_month: YYYYMM format
    tag: string like "30+DPD @ 75 DOB"
    """
    dpd, dob = parse_tag(tag)
    if dpd and dob:
        months_back = (dob // dpd) + 1
        run_date = datetime.strptime(run_month, '%Y%m')
        approved_date = run_date - relativedelta(months=months_back)
        return approved_date.strftime('%Y%m')
    return None

def calculate_psi_month(run_month):
    """
    Calculate PSI validation month (run month - 1)
    """
    run_date = datetime.strptime(run_month, '%Y%m')
    psi_date = run_date - relativedelta(months=1)
    return psi_date.strftime('%Y%m')

def is_date_format(value):
    """
    Check if a value is likely a date in YYYYMM format
    """
    try:
        str_value = str(value).strip()
        if '.' in str_value:
            str_value = str_value.split('.')[0] 
        if len(str_value) == 6 and str_value.isdigit():
            year = int(str_value[:4])
            month = int(str_value[4:])
            if 1900 <= year <= 2100 and 1 <= month <= 12:
                return True
    except:
        pass 
    return False

def convert_date_value(value):
    """
    Convert date from YYYYMM format to Mon-YY format
    Only converts values that are actually dates
    """
    try:
        if value is None:
            return value
        if isinstance(value, str):
            if is_date_format(value):
                date_obj = datetime.strptime(value.strip(), '%Y%m')
                return date_obj.strftime('%b-%y') 
        if isinstance(value, (int, float)):
            if is_date_format(int(value)):
                date_str = str(int(value))
                date_obj = datetime.strptime(date_str, '%Y%m')
                return date_obj.strftime('%b-%y')
        
        return value
        
    except:
        return value

def extract_model_and_segment(wb):
    """
    Extract model name and segment from the workbook
    """
    model_name = None
    segment = None
    
    first_sheet = wb[wb.sheetnames[0]]
    
    for row in range(1, min(11, first_sheet.max_row + 1)):
        for col in range(1, min(11, first_sheet.max_column + 1)):
            cell_value = first_sheet.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                
                if not model_name and any(keyword in cell_str.lower() for keyword in ['model', 'vehicle', 'car']):
                    if ':' in cell_str:
                        model_name = cell_str.split(':')[-1].strip()
                    else:
                        next_cell = first_sheet.cell(row=row, column=col+1).value
                        if next_cell:
                            model_name = str(next_cell).strip()
                
                if not segment and any(keyword in cell_str.lower() for keyword in ['segment', 'category', 'class']):
                    if ':' in cell_str:
                        segment = cell_str.split(':')[-1].strip()
                    else:
                        next_cell = first_sheet.cell(row=row, column=col+1).value
                        if next_cell:
                            segment = str(next_cell).strip()
    
    if model_name:
        model_name = re.sub(r'[^\w\s-]', '', model_name).strip()
        model_name = model_name.replace(' ', '_')
    
    if segment:
        segment = re.sub(r'[^\w\s-]', '', segment).strip()
        segment = segment.replace(' ', '_')
    
    return model_name, segment

def check_for_alerts(wb):
    """
    Check for 'red' or 'yellow' in Summary and Overall Comments fields
    Returns a dictionary with alert information
    """
    alerts = {
        'has_alerts': False,
        'summary': None,
        'overall_comments': None,
        'alert_details': []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name] 
        for row in range(1, min(50, ws.max_row + 1)):  # Search first 50 rows
            for col in range(1, min(10, ws.max_column + 1)):  # Search first 10 columns
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_str = str(cell.value).strip().lower()
                    
                    if 'summary' in cell_str and 'performance' not in cell_str:
                        for offset_col in range(1, 5):  # Check next 4 cells to the right
                            summary_cell = ws.cell(row=row, column=col + offset_col)
                            if summary_cell.value and str(summary_cell.value).strip():
                                summary_value = str(summary_cell.value).strip()
                                alerts['summary'] = summary_value
                                
                                if any(color in summary_value.lower() for color in ['red', 'yellow']):
                                    alerts['has_alerts'] = True
                                    alerts['alert_details'].append(f"Summary contains alert: {summary_value}")
                                break
                    
                    if 'overall' in cell_str and 'comment' in cell_str:
                        for offset_col in range(1, 5):  # Check next 4 cells to the right
                            comments_cell = ws.cell(row=row, column=col + offset_col)
                            if comments_cell.value and str(comments_cell.value).strip():
                                comments_value = str(comments_cell.value).strip()
                                alerts['overall_comments'] = comments_value
                                
                                if any(color in comments_value.lower() for color in ['red', 'yellow']):
                                    alerts['has_alerts'] = True
                                    alerts['alert_details'].append(f"Overall Comments contains alert: {comments_value}")
                                break
    
    return alerts

def process_excel_report_specific_cells(file_path, output_path=None, payload_model_name=None, payload_segment=None):
    """
    Read Excel file and convert only specific cells (B7, B14, B15, B17) from YYYYMM to Mon-YY format
    Also ensures B17 (Validation Performance Period) is the same as B14 (PSI Validation Period)
    Also checks for alerts in Summary and Overall Comments
    Additionally, copies vintage from Accuracy sheet (Benchmark row, Vintage column) to Overview B15 and B18 if they contain 'First production month'
    """
    wb = load_workbook(file_path, data_only=False)
    model_name, segment = extract_model_and_segment(wb)
    
    alerts = check_for_alerts(wb)
    
    vintage_value = None
    if 'Accuracy' in wb.sheetnames:
        accuracy_sheet = wb['Accuracy']
        try:
            benchmark_row = None
            vintage_col = None
            
            for row in range(1, min(20, accuracy_sheet.max_row + 1)):
                for col in range(1, min(20, accuracy_sheet.max_column + 1)):
                    cell_value = accuracy_sheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip().lower()
                        if cell_str == 'benchmark' and benchmark_row is None:
                            benchmark_row = row
                        if 'vintage' in cell_str and vintage_col is None:
                            vintage_col = col
            
            if benchmark_row and vintage_col:
                vintage_cell = accuracy_sheet.cell(row=benchmark_row, column=vintage_col)
                if vintage_cell.value is not None:
                    vintage_value = convert_date_value(vintage_cell.value)
            else:
                st.warning(f"Could not find Benchmark row or Vintage column in Accuracy sheet")
                
        except Exception as e:
            st.warning(f"Could not get vintage from Accuracy sheet: {e}")
    
    target_cells = ['B7', 'B14', 'B15', 'B16', 'B17']
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        psi_validation_value = None
        try:
            b14_cell = ws['B14']
            if b14_cell.value is not None:
                psi_validation_value = convert_date_value(b14_cell.value)
                b14_cell.value = psi_validation_value
        except Exception as e:
            st.warning(f"Could not process PSI Validation Period (B14): {e}")
        
        for cell_ref in target_cells:
            try:
                cell = ws[cell_ref]
                
                if cell_ref == 'B17' and psi_validation_value is not None:
                    old_value = cell.value
                    cell.value = psi_validation_value
                
                elif cell_ref == 'B14':
                    continue
                
                elif cell.value is not None:
                    converted_value = convert_date_value(cell.value)
                    
                    if converted_value != cell.value:
                        cell.value = converted_value                        
            except Exception as e:
                pass  # Silently skip cells that don't exist
        
        if sheet_name.lower() == 'overview' and vintage_value is not None:
            try:
                b15_cell = ws['B15']
                if b15_cell.value and str(b15_cell.value).strip().lower() == 'first production month':
                    b15_cell.value = vintage_value
            except Exception as e:
                st.warning(f"Could not process Overview B15: {e}")
            
            try:
                b18_cell = ws['B18']
                if b18_cell.value and str(b18_cell.value).strip().lower() == 'first production month':
                    b18_cell.value = vintage_value
            except Exception as e:
                st.warning(f"Could not process Overview B18: {e}")
    
    if output_path is None:
        if payload_model_name and payload_segment:
            clean_segment = payload_segment.strip()
            clean_segment = clean_segment.replace(':', '_')
            clean_segment = re.sub(r'[^\w\s-]', '', clean_segment)
            clean_segment = clean_segment.replace(' ', '_')
            output_filename = f"{payload_model_name}_{clean_segment}.xlsx"
        elif model_name and segment:
            output_filename = f"{model_name}_{segment}.xlsx"
        elif model_name:
            output_filename = f"{model_name}_unknown_segment.xlsx"
        else:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{base_name}_formatted.xlsx"
        
        output_path = os.path.join(processed_folder, output_filename)
    
    wb.save(output_path)
    
    return output_path, alerts

def download_and_process_report(payload, model_config):
    """
    Download a report with given parameters and process it
    """
    res = requests.get(
        'https://gds.paypalinc.com/qmonitor-api/excel_report/download/',    
        params=payload,  
        verify=True
    )
    
    if res.status_code == 200:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        model_name = model_config.get('model_name', 'unknown')
        segment = model_config.get('segment', 'unknown')
        filename = f"{model_name}_{segment}_{timestamp}.xlsx"
        filepath = os.path.join(reports_folder, filename)
        
        with open(filepath, 'wb') as file:
            file.write(res.content)
        
        try:
            processed_path, alerts = process_excel_report_specific_cells(
                filepath, 
                payload_model_name=payload.get('model_name'),
                payload_segment=payload.get('segment')
            ) 
            
            return filepath, processed_path, alerts
        except Exception as e:
            st.error(f"Error processing report: {str(e)}")
            return filepath, None, None
    else:
        st.error(f"Failed to download report. Status code: {res.status_code}")
        st.error(f"Response: {res.text}")
        return None, None, None

def get_current_run_month():
    """
    Get current month in YYYYMM format
    """
    return datetime.now().strftime('%Y%m')

def build_payload(model_config, run_month=None):
    if run_month is None:
        run_month = get_current_run_month()
    
    tag = model_config.get('tag', '')
    approved_month = calculate_approved_month(run_month, tag)
    psi_month = calculate_psi_month(run_month)
    
    payload = {
        'receiver': model_config.get('receiver', 'riybansal@paypal.com'),
        'model_name': model_config.get('model_name'),
        'segment': model_config.get('segment'),  
        'tag1': model_config.get('tag1', ''), 
        'mob1': model_config.get('mob1', '3'),
        'approved_month1': approved_month,
        'psi_month': psi_month
    }
    
    for key, value in model_config.items():
        if key not in ['model_name', 'segment', 'segment_prefix', 'tag', 'tag1', 'mob1']:
            if key not in payload:
                payload[key] = value
    
    return payload

MODEL_CONFIGS = {
    'Spyder3': {  
        'model_name': 'Spyder3', 
        'segments': [
            {'segment': 'DE_12M:Overall', 'product': '12m', 'tag': '30+DPD @ 180 DOB'},
            {'segment': 'DE_24M:Overall', 'product': '24m', 'tag': '30+DPD @ 180 DOB'},
            {'segment': 'DE_3M:Overall', 'product': '3m', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'DE_6M:Overall', 'product': '6m', 'tag': '30+DPD @ 180 DOB'},
            {'segment': 'DE_PI30:Overall', 'product': '30 days', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'ES:Overall', 'product': 'Pi3 (monthly)', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'IT:Overall', 'product': 'Pi3 (monthly)', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'AU:Overall', 'product': 'Pi4 (fortnightly)', 'tag': '30+DPD @ 75 DOB'},
            {'segment': 'FR:Overall', 'product': 'Pi4 (monthly)', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'GB:Overall', 'product': 'Pi3 (monthly)', 'tag': '30+DPD @ 90 DOB'},
        ], 
        'tag1': 'model_bad_tag'
    }, 

    'NTPP2': {
        'model_name': 'NTPP2',
        'segments': [
            {'segment': 'FR:Overall', 'product': 'Pi4 (monthly)', 'tag': '30+DPD @ 120 DOB'},
            {'segment': 'DE:Overall', 'product': 'Pi30, ST, LT', 'tag': '30+DPD @ 90 DOB'},
            {'segment': 'AU:Overall', 'product': 'Pi4 (fortnightly)', 'tag': '28+DPD @ 75 DOB'},
        ], 
        'tag1': 'approved'
    },
    
    'us_gpl_v2_rmr2_bq': {
        'model_name': 'us_gpl_v2_rmr2_bq',
        'segments': [
            {'segment': 'Overall', 'product': 'US GPL', 'tag': '10+DPD @ 30 DOB '},
        ],
        'tag1': 'model_bad_tag'
    }
}

def process_model(model_key, run_month=None):
    """
    Process a single model
    """
    if run_month is None:
        run_month = get_current_run_month()
        
    if model_key not in MODEL_CONFIGS:
        st.error(f"Error: Model '{model_key}' not found in MODEL_CONFIGS")
        st.error(f"Available models: {', '.join(MODEL_CONFIGS.keys())}")
        return []
    
    model_data = MODEL_CONFIGS[model_key]
    
    all_alerts = []
    
    # Create a progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_segments = len(model_data['segments'])
    
    for idx, segment_config in enumerate(model_data['segments']):
        config = {
            'model_name': model_data['model_name'],
            'segment': segment_config['segment'],
            'tag': segment_config['tag'],
            'product': segment_config.get('product', ''),
            'tag1': model_data.get('tag1', ''),
            'receiver': model_data.get('receiver', 'riybansal@paypal.com')
        }
        
        status_text.text(f"Processing: {model_data['model_name']} - {segment_config['segment']}")
        
        payload = build_payload(config, run_month)
        filepath, processed_path, alerts = download_and_process_report(payload, config)
        
        if processed_path:
            st.session_state.processed_reports.append({
                'model': model_data['model_name'],
                'segment': segment_config['segment'],
                'original_path': filepath,
                'processed_path': processed_path,
                'alerts': alerts
            })
        
        if alerts and alerts['has_alerts']:
            all_alerts.append({
                'model': model_data['model_name'],
                'segment': segment_config['segment'],
                'alerts': alerts
            })
        
        # Update progress
        progress_bar.progress((idx + 1) / total_segments)
        
        time.sleep(2)
    
    progress_bar.empty()
    status_text.empty()
    
    return all_alerts

# Streamlit UI
def main():
    st.title("📊 Report Processor")
    st.markdown("---")
    
    # Sidebar for configuration
    with st.sidebar:
        st.header("Configuration")
        
        # Model selection
        model_options = ["All Models"] + list(MODEL_CONFIGS.keys())
        selected_model = st.selectbox("Select Model", model_options)
        
        # Run month selection
        col1, col2 = st.columns(2)
        with col1:
            run_year = st.number_input("Year", min_value=2020, max_value=2030, value=datetime.now().year)
        with col2:
            run_month_num = st.number_input("Month", min_value=1, max_value=12, value=datetime.now().month)
        
        run_month = f"{run_year}{run_month_num:02d}"
        st.info(f"Run month: {run_month} ({datetime.strptime(run_month, '%Y%m').strftime('%b-%y')})")
        
        # Process button
        process_button = st.button("🚀 Process Reports", type="primary", use_container_width=True)
    
    # Main content area
    if process_button:
        st.session_state.processed_reports = []
        st.session_state.all_alerts = []
        
        with st.spinner("Processing reports..."):
            if selected_model == "All Models":
                for model_key in MODEL_CONFIGS.keys():
                    st.subheader(f"Processing {model_key}")
                    alerts = process_model(model_key, run_month)
                    if alerts:
                        st.session_state.all_alerts.extend(alerts)
            else:
                st.subheader(f"Processing {selected_model}")
                alerts = process_model(selected_model, run_month)
                if alerts:
                    st.session_state.all_alerts.extend(alerts)
        
        st.success("✅ Processing complete!")
    
    # Display results
    if st.session_state.processed_reports:
        st.header("📋 Processed Reports")
        
        # Create tabs for different views
        tab1, tab2, tab3 = st.tabs(["All Reports", "Alerts Only", "Download Links"])
        
        with tab1:
            for report in st.session_state.processed_reports:
                with st.expander(f"{report['model']} - {report['segment']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Original file:** {os.path.basename(report['original_path'])}")
                        st.write(f"**Processed file:** {os.path.basename(report['processed_path'])}")
                    with col2:
                        if report['alerts'] and report['alerts']['has_alerts']:
                            st.error("⚠️ Alert Found!")
                            st.write(f"**Summary:** {report['alerts']['summary']}")
                            st.write(f"**Comments:** {report['alerts']['overall_comments']}")
                        else:
                            st.success("✅ No alerts")
        
        with tab2:
            if st.session_state.all_alerts:
                st.error(f"⚠️ Found {len(st.session_state.all_alerts)} alerts!")
                for alert_info in st.session_state.all_alerts:
                    with st.expander(f"{alert_info['model']} - {alert_info['segment']}", expanded=True):
                        st.write(f"**Summary:** {alert_info['alerts']['summary']}")
                        st.write(f"**Overall Comments:** {alert_info['alerts']['overall_comments']}")
                        for detail in alert_info['alerts']['alert_details']:
                            st.write(f"- {detail}")
            else:
                st.success("✅ No alerts found in any reports")
        
        with tab3:
            st.write("### Download Processed Reports")
            for report in st.session_state.processed_reports:
                if report['processed_path'] and os.path.exists(report['processed_path']):
                    with open(report['processed_path'], 'rb') as file:
                        file_data = file.read()
                        filename = os.path.basename(report['processed_path'])
                        st.download_button(
                            label=f"📥 {report['model']} - {report['segment']}",
                            data=file_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
    
    # Footer
    st.markdown("---")
    st.markdown(f"Report folders: `{reports_folder}/` (original) | `{processed_folder}/` (processed)")

if __name__ == "__main__":
    main()